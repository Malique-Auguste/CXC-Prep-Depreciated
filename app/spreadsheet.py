import openpyxl

class Handler:
    def __init__(self, path):
        self.wb = openpyxl.load_workbook(path)

    def get_sheet_names(self):
        return self.wb.sheetnames

    def get_topics(self, subject):
        sheet = self.wb["Syllabus"]
        topics = []
        
        row_index = 3
        column_index = 0
        for column in sheet.iter_cols():
            column_index += 1
            if column[0].value == subject:
                break
        
        while sheet.cell(row = row_index, column = column_index).value.__str__() != "None":
            topics.append(Topic(sheet.cell(row = row_index, column = column_index).value.__str__(), sheet.cell(row = row_index, column = column_index + 1).value.__str__()))
            row_index += 1

        return topics

    def get_topic_probability(self, sheet_name):
        topics = self.get_topics(sheet_name)
        questions = self.get_questions(sheet_name)

        for topic in topics:
            for question in questions:
                if topic.code in question.topic:
                    topic.count += 1
        
        return topics



    def get_questions(self, sheet_name):
        sheet = self.wb[sheet_name]
        questions = []

        for row in sheet.iter_rows(min_row = 2):
            #print(row)
            questions.append(Question(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value))
        
        return questions
    
    def insert_html(self, file_path, data_type, text):
        file = open(file_path, "r")
        file_data = file.read()
        file.close()

        start_substring = f"<!--{data_type} starts here-->"
        s_s_index = file_data.find(start_substring)

        end_substring = f"<!--{data_type} ends here-->"
        e_s_index = file_data.find(end_substring)
        
        file_data = file_data[:s_s_index + len(start_substring) + 1] + text + file_data[e_s_index - 1:]

        file = open(file_path, "r+")
        file.seek(0)
        file.write(file_data)
        file.truncate()
        file.close()

        print("insert_html complete")

    def insert_js(self, file_path, data_type, text):
        file = open(file_path, "r")
        file_data = file.read()
        file.close()

        start_substring = f"/*starts here*/"
        s_s_index = file_data.find(start_substring)

        end_substring = f"/*ends here*/"
        e_s_index = file_data.find(end_substring)
        
        file_data = file_data[:s_s_index + len(start_substring) + 1] + text + file_data[e_s_index - 1:]

        file = open(file_path, "r+")
        file.seek(0)
        file.write(file_data)
        file.truncate()
        file.close()

        print("insert_js complete")



class Question:
    def __init__(self, year, january, paper, number, letters, topic):
        self.year = int(year)
        self.january = january
        self.paper = int(paper)
        self.number = int(number)
        self.letters = letters
        self.topic = topic

        self.html_text = ""
        self.str = ""

    def html(self):
        WITH_TABS = True

        if self.html_text == "":
            if WITH_TABS:
                self.html_text = "\t\t\t\t" + "<li data-topic = \"" + self.topic + "\">" + self.__str__() + "</li>"
            else:
                self.html_text = "<li data-topic = \"" + self.topic + "\">" + self.__str__() + "</li>"

        return self.html_text
    
    def __str__(self) -> str:
        if self.str == "":
            month = ""
            if self.january == 1:
                month = "January"
            else:
                month = "May/June"
            
            number = ""
            letters = self.letters.split("-")
            for i in range(0, len(letters)):
                if i > 0:
                    number += " - "
                number += '#' + self.number.__str__() + '.' + letters[i]
            
            self.str = month + " Paper " + self.paper.__str__() + ' ' + self.year.__str__() + " " + number
        
        return self.str

class Topic:
    def __init__(self, code, topic):
        self.code = code
        self.topic = topic
        self.count = 0
        self.html_text = ""
        self.js_data_text = ""
    
    def html(self):
        if self.html_text == "":
            self.html_text = "\t\t\t\t" + "<label><input type=\"checkbox\" class=\"topic\", value = \"" + self.code + "\" onclick=\"sort()\" checked = \"true\"><span>" + self.topic + "</span></label>"
         
        return self.html_text